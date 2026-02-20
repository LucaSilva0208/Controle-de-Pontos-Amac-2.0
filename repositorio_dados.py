import json
import os
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Arquivos
ARQUIVO_USUARIOS = "usuarios_sistema.json"
# Liste aqui todos os arquivos Excel que devem ser lidos
ARQUIVOS_EXCEL = ["funcionarios_sede.xlsx", "funcionarios_PGPL.xlsx"]
ARQUIVO_AFASTAMENTOS = "afastamentos.json"

class RepositorioDados:
    def __init__(self):
        # Garante que o arquivo de usuários existe com o admin padrão
        if not os.path.exists(ARQUIVO_USUARIOS):
            dados_iniciais = {
                "admin": {
                    "senha": "1234",
                    "perfil": "admin",
                    "unidade": "Sede Administrativa"
                }
            }
            with open(ARQUIVO_USUARIOS, 'w', encoding='utf-8') as f:
                json.dump(dados_iniciais, f, indent=4)

        # Garante a existência pelo menos do primeiro Excel (Principal)
        if openpyxl:
            principal = ARQUIVOS_EXCEL[0]
            if not os.path.exists(principal):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Funcionarios"
                # Cabeçalhos
                ws.append(["Nome Completo", "Matrícula", "Cargo", "Unidade", "Escala", "Carga Horária"])
                wb.save(principal)

        # Garante a existência do JSON de Afastamentos
        if not os.path.exists(ARQUIVO_AFASTAMENTOS):
            dados_afast = {
                "ferias": [],   # Lista de {matricula, inicio, fim}
                "recessos": []  # Lista de strings "YYYY-MM-DD"
            }
            with open(ARQUIVO_AFASTAMENTOS, 'w', encoding='utf-8') as f:
                json.dump(dados_afast, f, indent=4)

    # --- MÉTODOS DE USUÁRIOS (LOGIN) ---
    
    def buscar_usuario_login(self, login):
        """Lê o usuário do JSON"""
        try:
            with open(ARQUIVO_USUARIOS, 'r', encoding='utf-8') as f:
                dados = json.load(f)
                usuario = dados.get(login)
                if usuario:
                    usuario['login'] = login # Adiciona o login ao objeto
                    return usuario
        except Exception as e:
            print(f"Erro leitura: {e}")
        return None

    def salvar_novo_usuario(self, login, senha, perfil, unidade):
        """Escreve um novo usuário no JSON"""
        try:
            # 1. Lê tudo que já existe
            with open(ARQUIVO_USUARIOS, 'r', encoding='utf-8') as f:
                dados = json.load(f)
            
            # 2. Verifica se já existe
            if login in dados:
                return False, "Usuário já existe."

            # 3. Adiciona o novo
            dados[login] = {
                "senha": senha,
                "perfil": perfil,
                "unidade": unidade
            }

            # 4. Salva de volta no arquivo
            with open(ARQUIVO_USUARIOS, 'w', encoding='utf-8') as f:
                json.dump(dados, f, indent=4)
            
            return True, "Usuário criado com sucesso!"
        except Exception as e:
            return False, f"Erro ao salvar: {e}"

    def excluir_usuario(self, login):
        """Remove usuário do JSON"""
        if login == "admin": 
            return False, "Não pode excluir o Admin."
            
        try:
            with open(ARQUIVO_USUARIOS, 'r', encoding='utf-8') as f:
                dados = json.load(f)
            
            if login in dados:
                del dados[login]
                
                with open(ARQUIVO_USUARIOS, 'w', encoding='utf-8') as f:
                    json.dump(dados, f, indent=4)
                return True, "Usuário removido."
            else:
                return False, "Usuário não encontrado."
        except Exception as e:
            return False, f"Erro: {e}"

    def listar_usuarios(self):
        """Retorna lista para a tela de gestão"""
        try:
            with open(ARQUIVO_USUARIOS, 'r', encoding='utf-8') as f:
                dados = json.load(f)
                lista = []
                for login, infos in dados.items():
                    # Formata para o TreeView
                    lista.append((login, login, infos['unidade'], infos['perfil']))
                return lista
        except:
            return []

    # --- MÉTODOS DE FUNCIONÁRIOS (FOLHA) ---
    
    def listar_todos_funcionarios(self):
        """Lê do Excel funcionarios_sede.xlsx"""
        if not openpyxl: return []
        lista = []
        
        # Itera sobre todos os arquivos definidos na lista
        for arquivo in ARQUIVOS_EXCEL:
            if not os.path.exists(arquivo):
                continue
                
            try:
                wb = openpyxl.load_workbook(arquivo)
                
                # Define a unidade macro baseada no nome do arquivo
                nome_arquivo = os.path.basename(arquivo).lower()
                if "pgpl" in nome_arquivo:
                    unidade_macro = "PGPL"
                    carga_padrao = "30h"
                else:
                    unidade_macro = "Sede"
                    carga_padrao = "40h"

                # Itera sobre TODAS as abas (VIGIAS, ADM, etc)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Define escala padrão baseada no nome da aba
                    escala_aba = "NORMAL"
                    if "vigia" in sheet_name.lower() or "12x36" in sheet_name.lower():
                        escala_aba = "12X36"

                    # Ajuste fino de Carga Horária por Aba (ex: PGPL 40 HS vs PGPL 30 vs Prefeitura)
                    carga_aba = carga_padrao
                    sheet_lower = sheet_name.lower()
                    if "30" in sheet_lower or "prefeitura" in sheet_lower:
                        carga_aba = "30h"
                    elif "40" in sheet_lower:
                        carga_aba = "40h"

                    # Definição de colunas (Layout Padrão vs Prefeitura)
                    # Padrão: B(1)=Matrícula, C(2)=Nome, D(3)=Unidade, E(4)=Cargo
                    idx_mat, idx_nome, idx_unidade, idx_cargo = 1, 2, 3, 4
                    
                    # Layout Prefeitura (Deslocado): A(0)=Matrícula, B(1)=Nome, C(2)=Cargo
                    if "prefeitura" in sheet_lower:
                        idx_mat, idx_nome, idx_cargo = 0, 1, 2
                        idx_unidade = -1 # Não tenta ler unidade se não soubermos onde está neste layout
                    elif unidade_macro == "PGPL":
                        # Layout PGPL (Ajuste): A(0)=Matrícula, B(1)=Nome, C(2)=Unidade, D(3)=Cargo
                        idx_mat, idx_nome, idx_unidade, idx_cargo = 0, 1, 2, 3

                    # --- VARREDURA DE COLUNAS DE HORÁRIO (Busca Inteligente) ---
                    # Lê a primeira linha para mapear nomes -> índices
                    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                    
                    def encontrar_coluna(termos):
                        for i, h in enumerate(headers):
                            for t in termos:
                                if t in h: return i
                        return -1

                    idx_ent1 = encontrar_coluna(["ent. 1", "entrada 1", "ent 1", "1ª ent", "ent1"])
                    idx_sai1 = encontrar_coluna(["saída 1", "saida 1", "sai 1", "1ª sai", "sai1"])
                    idx_ent2 = encontrar_coluna(["ent. 2", "entrada 2", "ent 2", "2ª ent", "ent2"])
                    idx_sai2 = encontrar_coluna(["saída 2", "saida 2", "sai 2", "2ª sai", "sai2"])

                    # Pula o cabeçalho (min_row=2)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Verifica se tem Nome e Matrícula nos índices corretos
                        if len(row) > idx_nome and row[idx_nome] and row[idx_mat]: 
                            nome = str(row[idx_nome])
                            matricula = str(row[idx_mat])
                            
                            # Ignora linhas que sejam repetição de cabeçalho
                            if nome.strip().lower() == "nome" or matricula.strip().lower() == "matrícula":
                                continue
                            
                            cargo = ""
                            if len(row) > idx_cargo and row[idx_cargo]:
                                cargo = str(row[idx_cargo])
                            
                            # Reforço: Se o cargo for Vigia, força 12x36 independente da aba
                            escala_final = escala_aba
                            if "vigia" in cargo.lower():
                                escala_final = "12X36"
                            
                            # Detecção automática de status (Licença/Afastamento)
                            status = "Ativo"
                            texto_verificacao = (nome + " " + cargo).lower()
                            if "lic. sem vencimento" in texto_verificacao or "afastado" in texto_verificacao:
                                status = "Afastado"

                            # Lógica de Unidade (Lotação)
                            # Prioridade: O que está na célula do Excel > Unidade Macro do Arquivo
                            unidade_final = unidade_macro
                            if idx_unidade != -1 and len(row) > idx_unidade:
                                val_unidade = row[idx_unidade]
                                if val_unidade and str(val_unidade).strip():
                                    u_temp = str(val_unidade).strip()
                                    if u_temp.lower() != "lotação":
                                        unidade_final = u_temp

                            # Leitura dos Horários (Baseado na varredura)
                            ent1 = str(row[idx_ent1]) if idx_ent1 != -1 and len(row) > idx_ent1 and row[idx_ent1] else ""
                            sai1 = str(row[idx_sai1]) if idx_sai1 != -1 and len(row) > idx_sai1 and row[idx_sai1] else ""
                            ent2 = str(row[idx_ent2]) if idx_ent2 != -1 and len(row) > idx_ent2 and row[idx_ent2] else ""
                            sai2 = str(row[idx_sai2]) if idx_sai2 != -1 and len(row) > idx_sai2 and row[idx_sai2] else ""

                            lista.append({
                                "nome": nome,
                                "matricula": matricula,
                                "cargo": cargo,
                                "unidade": unidade_final,
                                "escala": escala_final, 
                                "carga_horaria": carga_aba,
                                "horario_ent1": ent1, "horario_sai1": sai1,
                                "horario_ent2": ent2, "horario_sai2": sai2,
                                "status": status
                            })
            except Exception as e:
                print(f"Erro ao ler Excel {arquivo}: {e}")
        
        return lista

    def adicionar_funcionario(self, nome, matricula, cargo, escala, carga_horaria="40h", ent1="", sai1="", ent2="", sai2=""):
        """Adiciona nova linha no Excel Principal (o primeiro da lista)"""
        if not openpyxl: return False, "Biblioteca openpyxl não instalada."
        try:
            # Por padrão, adiciona novos funcionários no primeiro arquivo da lista
            arquivo_alvo = ARQUIVOS_EXCEL[0]
            wb = openpyxl.load_workbook(arquivo_alvo)
            ws = wb.active
            
            # Adiciona seguindo o padrão detectado:
            # [Vazio, Matrícula, Nome, Unidade(Sede), Cargo, Escala, Carga, Ent1, Sai1, Ent2, Sai2]
            ws.append([None, matricula, nome, "Sede", cargo, escala, carga_horaria, ent1, sai1, ent2, sai2])
            
            wb.save(arquivo_alvo)
            return True, "Funcionário cadastrado com sucesso!"
        except Exception as e:
            return False, f"Erro ao salvar no Excel: {e}"

    def editar_funcionario(self, matricula_original, nome, matricula, cargo, escala, carga_horaria, ent1, sai1, ent2, sai2):
        """Edita funcionário procurando em TODOS os arquivos Excel"""
        if not openpyxl: return False, "Biblioteca openpyxl não instalada."
        for arquivo in ARQUIVOS_EXCEL:
            if not os.path.exists(arquivo): continue
            
            try:
                wb = openpyxl.load_workbook(arquivo)
                found = False
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Itera sobre as linhas (pula cabeçalho)
                    for row in ws.iter_rows(min_row=2):
                        # Layout: 0=None, 1=Matrícula, 2=Nome, 3=Unidade, 4=Cargo
                        if len(row) > 1 and row[1].value and str(row[1].value) == str(matricula_original):
                            row[1].value = matricula
                            row[2].value = nome
                            # row[3] Unidade mantém
                            if len(row) > 4: row[4].value = cargo
                            # Atualiza ou cria as células de horário
                            if len(row) > 7: row[7].value = ent1
                            if len(row) > 8: row[8].value = sai1
                            if len(row) > 9: row[9].value = ent2
                            if len(row) > 10: row[10].value = sai2
                            
                            found = True
                            break
                    if found: break
                
                if found:
                    wb.save(arquivo)
                    return True, "Funcionário atualizado com sucesso!"
            except Exception as e:
                print(f"Erro ao tentar editar em {arquivo}: {e}")
                
        return False, "Funcionário não encontrado em nenhum arquivo."

    def excluir_funcionario(self, matricula):
        """Remove a linha do funcionário procurando em TODOS os arquivos"""
        if not openpyxl: return False, "Biblioteca openpyxl não instalada."
        for arquivo in ARQUIVOS_EXCEL:
            if not os.path.exists(arquivo): continue
            
            try:
                wb = openpyxl.load_workbook(arquivo)
                found = False
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    row_idx = 0
                    # Precisamos do índice da linha para deletar
                    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if len(row) > 1 and row[1].value and str(row[1].value) == str(matricula):
                            row_idx = i
                            found = True
                            break
                    
                    if found:
                        ws.delete_rows(row_idx)
                        wb.save(arquivo)
                        return True, "Funcionário excluído com sucesso!"
                
            except Exception as e:
                print(f"Erro ao tentar excluir de {arquivo}: {e}")

        return False, "Funcionário não encontrado em nenhum arquivo."

    def buscar_funcionarios_por_unidade(self, unidade):
        todos = self.listar_todos_funcionarios()
        if unidade == "Todas": return todos
        return [f for f in todos if f.get('unidade') == unidade]

    def buscar_unidades(self):
        todos = self.listar_todos_funcionarios()
        unidades = list(set(f.get('unidade', 'Indefinido') for f in todos))
        unidades.sort()
        return unidades

    # --- MÉTODOS DE AFASTAMENTOS (JSON) ---

    def get_afastamentos(self):
        try:
            with open(ARQUIVO_AFASTAMENTOS, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"ferias": [], "recessos": []}

    def salvar_afastamento(self, matricula, inicio, fim, tipo="Férias"):
        dados = self.get_afastamentos()
        
        dados['ferias'].append({
            "matricula": matricula,
            "inicio": inicio, # string YYYY-MM-DD
            "fim": fim,       # string YYYY-MM-DD
            "tipo": tipo
        })
        
        with open(ARQUIVO_AFASTAMENTOS, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=4)

    def salvar_recessos_globais(self, lista_datas):
        dados = self.get_afastamentos()
        dados['recessos'] = lista_datas # Sobrescreve ou adiciona? O requisito diz "naquele lote", mas persiste. Vamos sobrescrever o atual.
        with open(ARQUIVO_AFASTAMENTOS, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=4)