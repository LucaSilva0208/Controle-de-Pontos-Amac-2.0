import os
import openpyxl

# Pega a pasta onde este script está
PASTA_BASE = os.path.dirname(os.path.abspath(__file__))

arquivos_para_verificar = [
    "funcionarios_sede.xlsx",
    "funcionarios_PGPL.xlsx"
]

print(f"--- DIAGNÓSTICO DE ARQUIVOS EXCEL ---")
print(f"Pasta do Sistema: {PASTA_BASE}\n")

total_funcionarios = 0

for nome_arq in arquivos_para_verificar:
    caminho_completo = os.path.join(PASTA_BASE, nome_arq)
    print(f"Verificando: {nome_arq} ...")
    
    if os.path.exists(caminho_completo):
        print(f"  [OK] Arquivo encontrado.")
        try:
            wb = openpyxl.load_workbook(caminho_completo)
            print(f"  -> Abas encontradas: {wb.sheetnames}")
            ws = wb.active
            
            # Conta linhas com dados (pulando cabeçalho)
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 2 and row[2]: # Verifica se tem Nome na coluna C (índice 2)
                    count += 1
            
            print(f"  [OK] Leitura bem sucedida.")
            print(f"  -> Funcionários encontrados: {count}")
            total_funcionarios += count
            
           
            # Verifica cabeçalhos
            headers = [cell.value for cell in ws[1]]
            print(f"  -> Colunas detectadas: {headers}")
            
        except Exception as e:
            print(f"  [ERRO] Falha ao ler o Excel: {e}")
    else:
        print(f"  [FALHA] Arquivo NÃO encontrado nesta pasta.")
        print(f"  -> O sistema esperava encontrar em: {caminho_completo}")
    print("-" * 30)

print(f"\nTOTAL GERAL DE FUNCIONÁRIOS NO SISTEMA: {total_funcionarios}")
input("\nPressione ENTER para sair...")
